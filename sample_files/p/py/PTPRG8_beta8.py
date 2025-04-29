import os
import json
import csv
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import pandas as pd
save_directory_path="./"

csv.field_size_limit(100000000)
output_dictionary ={}
standard_reports_label_array=[]
format_reports_label_array=[]
excel_file = 'keys.xlsx'
dfs = pd.read_excel(excel_file, sheet_name=None)
sheets = load_workbook(excel_file, read_only=True).sheetnames

def fix_nulls(s):
    for line in s:
        yield line.replace('\0', '')

def as_text(value):
    if value is None:
        return ""
    return str(value)

class GUI:

    def __init__(self, master):
        self.master = master
        master.title("Post Tika Processing Report Generator")
        self.directory_path = ""

        # Create the standard reports checkboxes
        standard_reports_frame = tk.LabelFrame(master, text="Standard Reports")
        standard_reports_frame.grid(row=1, column=0, padx=10, pady=10, sticky=tk.N)

        for i in range(0,4):
            standard_reports_label_array.append(sheets[i])
        standard_reports_labels = standard_reports_label_array

        self.standard_reports_vars = {}
        for i, label in enumerate(standard_reports_labels):
            var = tk.BooleanVar()
            cb = tk.Checkbutton(standard_reports_frame, text=label, variable=var)
            cb.grid(row=i//2, column=i%2, sticky="w")
            self.standard_reports_vars[label.lower().replace(" ", "_")] = var

        # Create the format reports checkboxes
        format_reports_frame = tk.LabelFrame(master, text="Format Reports")
        format_reports_frame.grid(row=1, column=2, padx=10, pady=10)

        for i in range(4,(len(sheets)-2)):
            format_reports_label_array.append(sheets[i]+ " Files")
        format_reports_labels = format_reports_label_array

        self.format_reports_vars = {}
        for i, label in enumerate(format_reports_labels):
            var = tk.BooleanVar()
            cb = tk.Checkbutton(format_reports_frame, text=label, variable=var)
            cb.grid(row=i//3, column=i%3, sticky="w")
            self.format_reports_vars[label.lower().replace(" ", "_")] = var

        # Create the buttons

        self.browse_button = tk.Button(master, text="Browse...", command=self.browse_directory)
        self.browse_button.grid(row=0, column=0)

        self.save_button = tk.Button(master, text="Save Location", command=self.choose_directory)
        self.save_button.grid(row=0, column=1)
        
        self.run_button = tk.Button(master, text="Run", command=self.process_directory)
        self.run_button.grid(row=0, column=2)
        
        self.cancel_button = tk.Button(master, text="Quit", command=root.destroy)
        self.cancel_button.grid(row=0, column=3)

        # Create the output window
        self.output = tk.scrolledtext.ScrolledText(master, height=50, width=130)
        self.output.grid(row=3, columnspan=4)

    def browse_directory(self):
        self.directory_path = filedialog.askdirectory()
        self.output.insert(tk.END, f"Selected processing directory: {self.directory_path}\n")

    def choose_directory(self):
        self.save_directory_path = filedialog.askdirectory()
        if self.save_directory_path:
            self.output.insert(tk.END, f"Selected save directory: {self.save_directory_path}\n")
            global save_directory_path
            save_directory_path=self.save_directory_path
            
    def process_directory(self):
        self.output.delete("1.0", tk.END)
        if not self.directory_path:
            self.output.insert(tk.END, "Please select a directory first.\n")
            return

        outer_resourcename2=""
        number_of_embedded_files=0
        risk_dict2=dict(zip(dfs['Risks']['fmt'].dropna(), dfs['Risks']['risk'].dropna()))

        if self.format_reports_vars["pdf_files"].get():
            pdf_keys = dfs['PDF']['key'].dropna().unique()
            pdf_csv_file = open("TIKA_temp_pdf.csv", "w", newline="", encoding='utf-8')
            pdf_csv_writer = csv.writer(pdf_csv_file,escapechar='\\')
            pdf_csv_writer.writerow(pdf_keys)

        if self.format_reports_vars["office_files"].get():
            office_keys, office_content = dfs['Office']['key'].dropna().unique(), dfs['Office']['mime'].dropna().unique()
            office_csv_file = open("TIKA_temp_office.csv", "w", newline="", encoding='utf-8')
            office_csv_writer = csv.writer(office_csv_file,escapechar='\\')
            office_csv_writer.writerow(office_keys)

        if self.format_reports_vars["image_files"].get():
            image_keys = dfs['Image']['key'].dropna().unique()
            image_csv_file = open("TIKA_temp_image.csv", "w", newline="", encoding='utf-8')
            image_csv_writer = csv.writer(image_csv_file,escapechar='\\')
            image_csv_writer.writerow(image_keys)

        if self.format_reports_vars["audio_video_files"].get():
            av_keys = dfs['Audio Video']['key'].dropna().unique()
            av_csv_file = open("TIKA_temp_av.csv", "w", newline="", encoding='utf-8')
            av_csv_writer = csv.writer(av_csv_file,escapechar='\\')
            av_csv_writer.writerow(av_keys)

        if self.format_reports_vars["design_files"].get():
            design_keys, design_content = dfs['Design']['key'].dropna().unique(), dfs['Design']['mime'].dropna().unique().tolist()
            design_csv_file = open("TIKA_temp_design.csv", "w", newline="", encoding='utf-8')
            design_csv_writer = csv.writer(design_csv_file,escapechar='\\')
            design_csv_writer.writerow(design_keys)

        if self.format_reports_vars["web_files"].get():
            web_keys, web_content = dfs['Web']['key'].dropna().unique(), dfs['Web']['mime'].dropna().unique().tolist()
            web_csv_file = open("TIKA_temp_web.csv", "w", newline="", encoding='utf-8')
            web_csv_writer = csv.writer(web_csv_file,escapechar='\\')
            web_csv_writer.writerow(web_keys)

        if self.format_reports_vars["email_files"].get():
            email_keys, email_content = dfs['Email']['key'].dropna().unique(), dfs['Email']['mime'].dropna().unique().tolist()
            email_csv_file = open("TIKA_temp_email.csv", "w", newline="", encoding='utf-8')
            email_csv_writer = csv.writer(email_csv_file,escapechar='\\')
            email_csv_writer.writerow(email_keys)

        if self.format_reports_vars["gis_files"].get():
            gis_keys, gis_content = dfs['GIS']['key'].dropna().unique(), dfs['GIS']['mime'].dropna().unique().tolist()
            gis_csv_file = open("TIKA_temp_gis.csv", "w", newline="", encoding='utf-8')
            gis_csv_writer = csv.writer(gis_csv_file,escapechar='\\')
            gis_csv_writer.writerow(gis_keys)

        if self.format_reports_vars["data_files"].get():
            data_keys, data_content = dfs['Data']['key'].dropna().unique(), dfs['Data']['mime'].dropna().unique().tolist()
            data_csv_file = open("TIKA_temp_data.csv", "w", newline="", encoding='utf-8')
            data_csv_writer = csv.writer(data_csv_file,escapechar='\\')
            data_csv_writer.writerow(data_keys)

        if self.format_reports_vars["legacy_files"].get():
            legacy_keys, legacy_content = dfs['Legacy']['key'].dropna().unique(), dfs['Legacy']['mime'].dropna().unique().tolist()
            legacy_csv_file = open("TIKA_temp_legacy.csv", "w", newline="", encoding='utf-8')
            legacy_csv_writer = csv.writer(legacy_csv_file,escapechar='\\')
            legacy_csv_writer.writerow(legacy_keys)

        if self.standard_reports_vars["file_listing"].get():
            file_listing_keys = dfs['File Listing']['key'].dropna().unique()
            file_listing_csv_file = open("TIKA_temp_fl.csv", "w", newline="", encoding='utf-8')
            file_listing_csv_writer = csv.writer(file_listing_csv_file,escapechar='\\')
            file_listing_csv_writer.writerow(file_listing_keys)

        if self.standard_reports_vars["passwords_encryption"].get():
            pass_encrypt_keys = dfs['Passwords Encryption']['key'].dropna().unique()
            pw_csv_file = open("TIKA_temp_pw.csv", "w", newline="", encoding='utf-8')
            pw_csv_writer = csv.writer(pw_csv_file,escapechar='\\')
            pw_csv_writer.writerow(pass_encrypt_keys)

        if self.standard_reports_vars["compressed_files"].get():
            compressed_keys, compressed_content = dfs['Compressed Files']['key'].dropna().unique(), dfs['Compressed Files']['mime'].dropna().unique().tolist()
            compresssed_csv_file = open("TIKA_temp_compressed.csv", "w", newline="", encoding='utf-8')
            compresssed_csv_writer = csv.writer(compresssed_csv_file,escapechar='\\')
            compresssed_csv_writer.writerow(compressed_keys)

        if self.standard_reports_vars["zero_byte_data"].get():
            zb_keys = dfs['Zero Byte Data']['key'].dropna().unique()
            zb_csv_file = open("TIKA_temp_zb.csv", "w", newline="", encoding='utf-8')
            zb_csv_writer = csv.writer(zb_csv_file,escapechar='\\')
            zb_csv_writer.writerow(zb_keys)

        json_counter=0

        for root, dirs, files in os.walk(self.directory_path):

            files = sorted(files, key=str.casefold)
            json_files = [file for file in files if file.endswith(".json")]
            for file in json_files:

                self.output.insert(tk.END,"Processing " + file + "\n")
                self.output.see(tk.END)
                self.output.update()
                file_path = os.path.join(root, file)

                with open(file_path, errors="ignore") as f:
                    json_counter=json_counter+1
                    data = json.load(f)
                    counter=0
                    if isinstance(data, list):

                        if self.format_reports_vars["pdf_files"].get():

                            for item in data:

                                counter=counter+1
                                number_of_embedded_files=0
                                get_field=[]
                                content_type = item.get("Content-Type")
                                ResourceName = item.get("resourceName") if item.get("resourceName") else "NO FILE NAME GIVEN"
                                if counter==1:
                                    outer_ResourceName=item.get("resourceName")

                                X_TIKA_embedded_depth=item.get("X-TIKA:embedded_depth")

                                if content_type == "application/pdf"  and X_TIKA_embedded_depth == "0":
                                    number_of_embedded_files=len(data)
                                    outer_resourcename2= ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get("resourceName")))
                                    get_field.insert(0,outer_resourcename2)
                                    get_field.insert(1,"")

                                if content_type == "application/pdf" and X_TIKA_embedded_depth != "0" and counter > 1:
                                    number_of_embedded_files=0
                                    get_field.insert(1,outer_ResourceName)
                                    ResourceName2= ILLEGAL_CHARACTERS_RE.sub(r'', str(ResourceName))
                                    get_field.insert(0,ResourceName2)

                                for j in pdf_keys:
                                    if content_type == "application/pdf":
                                        raw=item.get(j)
                                        if j != "resourceName" and raw is not None:
                                            raw2 = ILLEGAL_CHARACTERS_RE.sub(r'', str(raw))
                                            get_field.append(raw2)

                                        else:
                                            if raw is None and j != "Embedded In" and j != "resourceName":
                                                get_field.append("")

                                if len(get_field) > 0:
                                    if number_of_embedded_files >=1:
                                        number_of_embedded_files=number_of_embedded_files-1
                                    get_field[4]=number_of_embedded_files
                                    pdf_csv_writer.writerow(get_field)
                                    output_dictionary.update({'pdf_files': [pdf_csv_file, 'PDF_Report.xlsx']})

                        if self.format_reports_vars["office_files"].get():
                            for item in data:
                                get_field = []
                                content_type = item.get("Content-Type")

                                if content_type and ';' in content_type:
                                    content_type = content_type.split(';', 1)[0]

                                if content_type in office_content:
                                    get_field = [ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get(i))) if item.get(i) is not None else "" for i in office_keys]

                                    if get_field:
                                        office_csv_writer.writerow(get_field)
                                        output_dictionary.update({'office_files': [office_csv_file, 'Office_Report.xlsx']})

                        if self.format_reports_vars["image_files"].get():
                            for item in data:
                                get_field = []
                                content_type = item.get("Content-Type")
                                
                                if content_type is not None and "image/" in content_type:
                                    resourceName = item.get("resourceName")
                                    get_field = [ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get(i))) if item.get(i) is not None else "" for i in image_keys]

                                    if resourceName is None:
                                        get_field[0] = "EMBEDDED FILE (NO FILE NAME)"

                                    if get_field:
                                        image_csv_writer.writerow(get_field)
                                        output_dictionary.update({'image_files': [image_csv_file, 'Image_Report.xlsx']})

                        if self.format_reports_vars["audio_video_files"].get():
                            for item in data:
                                get_field = []
                                content_type = item.get("Content-Type")

                                if content_type and ';' in content_type:
                                    content_type = content_type.split(';', 1)[0]

                                if content_type is not None and ('audio' in content_type or 'video' in content_type):
                                    get_field = [ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get(i))) if item.get(i) is not None else "" for i in av_keys]

                                    if get_field:
                                        av_csv_writer.writerow(get_field)
                                        output_dictionary.update({'audio_video_files': [av_csv_file, 'AV_Report.xlsx']})

                        if self.format_reports_vars["design_files"].get():
                            for item in data:
                                get_field = []
                                content_type = item.get("Content-Type")

                                if content_type and ';' in content_type:
                                    content_type = content_type.split(';', 1)[0]

                                if content_type in design_content:
                                    get_field = [ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get(i))) if item.get(i) is not None else "" for i in design_keys]

                                    if get_field:
                                        design_csv_writer.writerow(get_field)
                                        output_dictionary.update({'design_files': [design_csv_file, 'Design_Report.xlsx']})

                        if self.format_reports_vars["web_files"].get():
                            data_dict = dict(item for item in data[0].items())
                            resourceName=ILLEGAL_CHARACTERS_RE.sub(r'', str(data_dict['resourceName']))

                            if 'X-TIKA:Parsed-By-Full-Set' in data_dict.keys():
                                is_warc=ILLEGAL_CHARACTERS_RE.sub(r'', str(data_dict['X-TIKA:Parsed-By-Full-Set']))
                            else:
                                is_warc=""
                                
                            for item in data:
                                counter=2
                                content_type = item.get("Content-Type")
                                X_TIKA_embedded_depth=item.get("X-TIKA:embedded_depth")
                                get_field=[None]*len(web_keys)

                                if content_type and ';' in content_type:
                                    content_type = content_type.split(';', 1)[0]
                                    
                                if content_type in web_content or "WARC" in is_warc:
                                    if int(X_TIKA_embedded_depth) == 0:
                                        get_field[0]=resourceName
                                        number_of_embedded_files=len(data)

                                        for i in web_keys:
                                            if i != "resourceName" and i != "# Embedded Files":

                                                raw=item.get(i)
                                                if raw is not None:
                                                    raw2 = ILLEGAL_CHARACTERS_RE.sub(r'', str(raw))
                                                    get_field[counter]=raw2
                                                    counter=counter+1
                                                else:
                                                    get_field[counter]=raw
                                                    counter=counter+1

                                        if len(get_field) > 0:
                                            if number_of_embedded_files >=1:
                                                number_of_embedded_files=number_of_embedded_files-1
                                                get_field[1]=number_of_embedded_files
                                            web_csv_writer.writerow(get_field)
                                            output_dictionary.update({'web_files': [web_csv_file, 'Web_Report.xlsx']})

                        if self.format_reports_vars["email_files"].get():
                            data_dict = dict(item for item in data[0].items())
                            outer_resourcename2=ILLEGAL_CHARACTERS_RE.sub(r'', str(data_dict['resourceName']))
                            outer_content_type=ILLEGAL_CHARACTERS_RE.sub(r'', str(data_dict['Content-Type']))

                            for item in data:
                                resource_name=item.get('resourceName')
                                if resource_name == None:
                                    resource_name="EMBEDDED EMAIL"
                                content_type = item.get("Content-Type")
                                X_TIKA_embedded_depth=item.get("X-TIKA:embedded_depth")
                                get_field=[None]*len(email_keys)
                                if isinstance(content_type, list):
                                    content_type = content_type[0]
                                elif content_type and ';' in content_type:
                                    content_type = content_type.split(';', 1)[0]

                                if outer_content_type in email_content:
                                    get_field[0]=resource_name

                                    if X_TIKA_embedded_depth == '0':
                                        get_field[1]=""
                                        number_of_embedded_files=len(data)
                                    else:
                                        number_of_embedded_files=0
                                        get_field[1]=outer_resourcename2

                                    counter=3
                                    for i in email_keys:
                                        if i != "resourceName" and i != "Embedded In" and i != "# Emails":
                                            raw=item.get(i)

                                            if raw is not None:
                                                raw2 = ILLEGAL_CHARACTERS_RE.sub(r'', str(raw))
                                                get_field[counter]=raw2
                                                counter=counter+1

                                            else:
                                                get_field[counter]=raw
                                                counter=counter+1

                                    if len(get_field) > 0:
                                        if number_of_embedded_files >=1:
                                            get_field[2]=number_of_embedded_files-1
                                        email_csv_writer.writerow(get_field)
                                        output_dictionary.update({'email_files': [email_csv_file, 'Email_Report.xlsx']})

                        if self.format_reports_vars["gis_files"].get():
                            for item in data:
                                get_field = []
                                content_type = item.get("Content-Type")

                                if content_type and ';' in content_type:
                                    content_type = content_type.split(';', 1)[0]

                                if content_type in gis_content:
                                    get_field = [ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get(i))) if item.get(i) is not None else "" for i in gis_keys]

                                    if get_field:
                                        gis_csv_writer.writerow(get_field)
                                        output_dictionary.update({'gis_files': [gis_csv_file, 'GIS_Report.xlsx']})

                        if self.format_reports_vars["data_files"].get():
                            for item in data:
                                get_field = []
                                content_type = item.get("Content-Type")

                                if content_type in data_content:
                                    get_field = [ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get(i))) if item.get(i) is not None else "" for i in data_keys]

                                if len(get_field) > 0:
                                    data_csv_writer.writerow(get_field)
                                    output_dictionary.update({'data_files': [data_csv_file, 'Data_Report.xlsx']})

                        if self.format_reports_vars["legacy_files"].get():
                            for item in data:
                                get_field = [None] * len(legacy_keys)
                                content_type = item.get("Content-Type")

                                if content_type and ';' in content_type:
                                    content_type = content_type.split(';', 1)[0]

                                if content_type in legacy_content:
                                    resource_name = ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get('resourceName')))
                                    get_field[0] = resource_name if resource_name != "None" else "NO FILE NAME"

                                    for counter, i in enumerate(legacy_keys):
                                        if i != "resourceName":
                                            raw = ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get(i)))
                                            get_field[counter] = raw if raw != "None" else ""

                                if len(get_field) > 0 and all(element is not None for element in get_field):
                                    legacy_csv_writer.writerow(get_field)
                                    output_dictionary.update({'legacy_files': [legacy_csv_file, 'Legacy_Report.xlsx']})

                        if self.standard_reports_vars["file_listing"].get():
                            data_dict = dict(item for item in data[0].items())
                            outer_resourcename2=ILLEGAL_CHARACTERS_RE.sub(r'', str(data_dict['resourceName']))

                            for item in data:
                                get_field=[None]*len(file_listing_keys)
                                resource_name=ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get('resourceName')))
                                X_TIKA_embedded_depth=item.get("X-TIKA:embedded_depth")
                                get_field[0] = resource_name if resource_name != "None" else "EMBEDDED FILE (NO FILE NAME)"
                                counter=0

                                risk_check(item, risk_dict2)

                                if X_TIKA_embedded_depth == '0':
                                    get_field[1] = ""
                                    number_of_embedded_files = len(data)
                                else:
                                    warc_resourceName = item.get("warc:WARC-Target-URI")
                                    if warc_resourceName:
                                        get_field[0] = warc_resourceName if warc_resourceName else "EMBEDDED FILE (NO FILE NAME)"
                                    else:
                                        get_field[0] = resource_name if resource_name != "None" else "EMBEDDED FILE (NO FILE NAME)"
                                    number_of_embedded_files=0
                                    get_field[1]=outer_resourcename2

                                for j in file_listing_keys:
                                    raw=item.get(j)
                                    if raw is not None and j != "Embedded In" and j != "resourceName":
                                        raw2 = ILLEGAL_CHARACTERS_RE.sub(r'', str(raw))
                                        get_field[counter]=raw2
                                    counter=counter+1

                                if len(get_field) > 0:
                                    if number_of_embedded_files >=1:
                                        number_of_embedded_files=number_of_embedded_files-1
                                    get_field[2]=number_of_embedded_files
                                    get_field[3]=risk_val
                                    file_listing_csv_writer.writerow(get_field)
                                    output_dictionary.update({'file_listing': [file_listing_csv_file, 'File_Listing_Report.xlsx']})

                        if self.standard_reports_vars["passwords_encryption"].get():
                            for item in data:
                                x_tika_encrypted = item.get("X-TIKA:encrypted")
                                pdf_encrypted = item.get("pdf:encrypted")
                                get_field = [ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get(i))) if item.get(i) is not None else "" for i in pass_encrypt_keys]

                                if x_tika_encrypted == "true" or pdf_encrypted == "true":
                                    pw_csv_writer.writerow(get_field)
                                    output_dictionary.update({'passwords_encryption': [pw_csv_file, 'Password_Encryption_Report.xlsx']})

                        if self.standard_reports_vars["compressed_files"].get():
                            for item in data:
                                get_field = []
                                content_type = item.get("Content-Type")

                                if content_type in compressed_content:
                                    number_of_embedded_files = len(data) - 1
                                    for i in compressed_keys:
                                        raw = ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get(i)))
                                        get_field.append(raw if raw != "None" else "")

                                    if len(get_field) > 0:
                                        get_field[1] = number_of_embedded_files
                                        compresssed_csv_writer.writerow(get_field)
                                        output_dictionary.update({'compressed_files': [compresssed_csv_file, 'Compressed_Report.xlsx']})

                        if self.standard_reports_vars["zero_byte_data"].get():

                            for item in data:
                                X_TIKA_embedded_depth = item.get("X-TIKA:embedded_depth")
                                file_size = item.get("Content-Length")

                                if X_TIKA_embedded_depth == "0" and file_size == "0":
                                    get_field = [ILLEGAL_CHARACTERS_RE.sub(r'', str(item.get(j))) if item.get(j) else item.get(j) for j in zb_keys]

                                    if len(get_field) > 0:
                                        zb_csv_writer.writerow(get_field)
                                        output_dictionary.update({'zero_byte_data': [zb_csv_file, 'Zero_Byte_Report.xlsx']})

        self.output.insert(tk.END,"\nFinished processing " + str(json_counter) + " JSON files.\nWriting Excel files, please wait...")
        self.output.see(tk.END)
        self.output.update()

        report_array=["pdf_files","image_files","office_files","image_files","audio_video_files","design_files","web_files","email_files","gis_files","data_files","legacy_files"]
        report_array_standard=["file_listing","passwords_encryption","compressed_files","zero_byte_data"]
        all_report_vars = {**self.format_reports_vars, **self.standard_reports_vars}
        all_array = report_array + report_array_standard

        for i in all_array:
            if i in output_dictionary and all_report_vars[i].get():
                csvfile, outputfile = output_dictionary[i]
                clean_xls(csvfile, outputfile)

        try:
            prefixed = [filename for filename in os.listdir('.') if filename.startswith("TIKA_temp")]
            for i in prefixed:
                os.remove(i)
        except Exception:
            print("Problem deleting temp csv files, is one open in a program somewhere?")
            pass

        self.output.insert(tk.END,"\nDone.")
        self.output.see(tk.END)
        self.output.update()
        tk.messagebox.showinfo("Done", "Parsing complete")

def clean_xls(csv_file, output_file_name):
    csv_file.close()
    wb = Workbook()
    ws = wb.active
    header = NamedStyle(name="header1a")
    with open(csv_file.name, 'r', encoding='utf-8') as f:
        for row in csv.reader(fix_nulls(f)):
            ws.append(row)

        style_1 = Font(name='Calibri', size=12)
        header.font = Font(name='Calibri', size=12,bold=True)
        header_row = ws[1]

        for cell in ws['A'] + ws[1]:
            cell.font = style_1
            for cell in header_row:
                cell.style = header
        sheet = wb.active
        sheet.freeze_panes = 'A2'
        sheet.print_title_rows = '1:1'
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            new_column_length = max(len(as_text(cell.value)) for cell in column_cells)
            new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                sheet.column_dimensions[new_column_letter].width = new_column_length + 2
    if save_directory_path:
        wb.save(save_directory_path+os.sep+output_file_name)
    else:
        wb.save(output_file_name)

def risk_check(item, risk_dict2):
    global risk_val
    sf_pronom_id=item.get("sf:pronom:id")
    if sf_pronom_id in risk_dict2:
        risk_val=risk_dict2[sf_pronom_id]
    else:
        risk_val="UNKNOWN"
    return risk_val

root = tk.Tk()
gui = GUI(root)
root.mainloop()
