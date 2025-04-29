import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle
import openpyxl
from openpyxl import load_workbook
import shutil

if os.path.isfile("all.xlsx"):
    os.remove("all.xlsx")
shutil.copy("all2.xlsx","all.xlsx")
destfile="all.xlsx"

def as_text(value):
    if value is None:
        return ""
    return str(value)

counter=1
i = 974
df_all = pd.DataFrame()  
for f in ['1130','1268']:#,'1406','1544','1682','1820','1958','2096','2234','2372','2510','2648','2786','2924','3062','3200','3338','3476','3614','3752','3890','4028','4166','4304','4442','4580','4718','NONEA']:    
    df_name="df"+str(counter)
    print("DF NAME IS " + df_name)
    file="Collective_Bargaining_1995_RESTRICTED_"+f+".csv"
    i = i + 32
    counterj=29-counter
    print("OPENING FILE " + file + " to insert " + str(counterj) + " columns at " + str(i))

    if os.path.exists(file):#THIS WILL CAUSE PROBLEMS WHEN THERE ARE NO CSV FILES, INSERT FAKE EMPTY ONES TO FIX
       
        df1 = pd.read_csv(file, sep="|")
        
        for j in range(0,counterj):
            counter_value=str(j)
            
            df1 = pd.concat([df1.iloc[:, :i], pd.DataFrame('', columns=['SUBKEY'+str(j), 'D_ERR_CT'+str(j), 'AHEAMT'+str(j), 'AHEYEAR'+str(j), 'AHEMNTH'+str(j), 'AHEDAY'+str(j), 'REOPYR'+str(j), 'REOPMO'+str(j), 'REOPDY'+str(j), 'REOPTYPE'+str(j), 'NUMEMPLS'+str(j), 'EMPLYR'+str(j),'EMPLMO'+str(j), 'EMPLYDY'+str(j), 'WGEOVERL'+str(j),'WGEFORM'+str(j), 'WGEUNIF'+str(j), 'WGCTSCHG'+str(j), 'WGPCTCHG'+str(j), 'WAGEYR'+str(j), 'WAGEMO'+str(j), 'WAGEDY'+str(j), 'REASON'+str(j), 'ESCOVERL'+str(j), 'ESCPCCHG'+str(j), 'ESCCTCHG'+str(j), 'ESCYR'+str(j), 'ESCMO'+str(j), 'ESCDY'+str(j), 'DACTCODE'+str(j), 'DEFCHG'+str(j), 'CHGDATE'+str(j)], index=df1.index), df1.iloc[:, i:]], axis=1)
    print(df1)
    frames=[df1]
    df_all = pd.concat(frames)
    counter=counter+1
print("#################")
print(df_all)
# print("DONE LOOPING OVER CSV FILES, WRITE OUTPUT TO ALL.XLSX")
# with pd.ExcelWriter("all.xlsx", mode="a", engine="openpyxl") as writer:
#     df_all.to_excel(writer, sheet_name="Collective Bargaining", index=False)#save it to the 'new_tab' in destfile
# # 
# print("DONE writing worksheets, starting to fix sheet.")
# wb = load_workbook("all.xlsx")
# if 'Sheet1' in wb.sheetnames:
#     del wb['Sheet1']
# ws = wb.worksheets[0]
# #print(ws)
# header = NamedStyle(name="header")
# header.font = Font(name='Arial', size=11,bold=True)
# style_1 = Font(name='Arial', size=11)   
# header_row = ws[1]
# 
# for cell in ws['A'] + ws[1]:
#     cell.font = style_1
#     for cell in header_row:
#         cell.style = header
# ws.freeze_panes = 'A2'
# ws.print_title_rows = '1:1'
# ws.auto_filter.ref = ws.dimensions
# for column_cells in ws.columns:
#     new_column_length = max(len(as_text(cell.value)) for cell in column_cells)
#     new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
#     if new_column_length > 0:
#         ws.column_dimensions[new_column_letter].width = new_column_length + 5
# wb.save("all.xlsx")
# print("ALL DONE.")

