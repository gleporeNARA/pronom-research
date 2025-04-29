import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle
import openpyxl
from openpyxl import load_workbook


destfile="all.xlsx"

def as_text(value):
    if value is None:
        return ""
    return str(value)

for f in ['1130','1268','1406','1544','1682','1820','1958','2096','2234','2372','2510','2648','2786','2924','3062','3200','3338','3476','3614','3752','3890','4028','4166','4304','4442','4580','4718','NONEA']:
    
    file="Collective_Bargaining_1974_RESTRICTED_"+f+".xlsx"
    print(file)

    if os.path.exists(file):
       
        data = pd.read_excel(file, sheet_name="Sheet") #change xxx with the sheet name that includes the data
        with pd.ExcelWriter("all.xlsx", mode="a", engine="openpyxl") as writer:
            data.to_excel(writer, sheet_name=f, index=False)#save it to the 'new_tab' in destfile
        

print("DONE writing worksheets, starting to fix sheets.")
wb = load_workbook(filename = destfile)
del wb['Sheet1']
header = NamedStyle(name="header")
header.font = Font(name='Arial', size=11,bold=True)
style_1 = Font(name='Arial', size=11)        
        
for f in ['1130','1268','1406','1544','1682','1820','1958','2096','2234','2372','2510','2648','2786','2924','3062','3200','3338','3476','3614','3752','3890','4028','4166','4304','4442','4580','4718','NONEA']:
    
    if f in wb.sheetnames:
        ws = wb[f]
        print(ws)
        
        header_row = ws[1]
        
        for cell in ws['A'] + ws[1]:
            cell.font = style_1
            for cell in header_row:
                cell.style = header
        ws.freeze_panes = 'A2'
        ws.print_title_rows = '1:1'
        ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            new_column_length = max(len(as_text(cell.value)) for cell in column_cells)
            new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                ws.column_dimensions[new_column_letter].width = new_column_length + 5
wb.save('all_done.xlsx')
print("ALL DONE.")
