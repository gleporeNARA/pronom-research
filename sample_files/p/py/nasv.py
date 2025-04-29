#!/usr/bin/env python3
import binascii
import sys
from pandas import DataFrame
import pandas as pd
import numpy as np
from numpy import loadtxt
import codecs
from GLebcdic import Packed_Decimal_Clean, Zoned_Decimal, EBCDIC, Binary
import csv
import openpyxl
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle

def Extract(lst,element): 
    return [item[element] for item in lst] 

def as_text(value):
    if value is None:
        return ""
    return str(value)

filename="NASV"

path2 = sys.argv[1]

df1 = pd.read_excel("layout.xlsx") # can also index sheet by name or fetch all sheets
L = df1.values.tolist()

widths=Extract(L,1)
widths = [int(i)*2 for i in widths] 
colnames=Extract(L,0)
data_type=Extract(L,2)
    
df = pd.read_fwf(path2, widths=widths,names=colnames, dtype="string", engine = 'c')
#df.fillna("",inplace=True)####TEMP REMOVE
print(df[:10])

columns = list(df) 
count=0
for j in columns: 
    convert=data_type[count]
    print("Procesing column " + str(j) + " as " + str(convert))
    
    if convert=="EBCDIC":
        df[j]=df[j].apply(EBCDIC)
    if convert=="Packed":
        df[j]=df[j].apply(Packed_Decimal_Clean)
    if convert=="Zoned":
        df[j]=df[j].apply(Zoned_Decimal)
    if convert=="Binary":
        df[j]=df[j].apply(Binary)
    if convert=="Plain":
        print(df[j])
        df[j]=df[j]
    count=count+1
    
df.replace("+0","0", inplace=True)
df.replace("+","", inplace=True)
df.replace("-","", inplace=True)
print(df[:10])

df.to_csv(filename+r'.csv', index=False,sep = '|')

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)
   
style_1 = Font(name='Calibri', size=12)

header = NamedStyle(name="header")
header.font = Font(name='Calibri', size=12,bold=True)
header_row = ws[1]

for cell in ws['A'] + ws[1]:
    #cell.style = 'Pandas'
    cell.font = style_1
    for cell in header_row:
        cell.style = header
sheet = wb.active
sheet.freeze_panes = 'A2'
sheet.print_title_rows = '1:1'
sheet.auto_filter.ref = sheet.dimensions
for column_cells in sheet.columns:
    new_column_length = max(len(as_text(cell.value)) for cell in column_cells)
    new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
    if new_column_length > 0:
        sheet.column_dimensions[new_column_letter].width = new_column_length + 5
wb.save(filename+r'.xlsx')


